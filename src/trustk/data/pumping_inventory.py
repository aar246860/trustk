"""Inventory USGS Lovelock pumping-test records and drawdown products."""

from __future__ import annotations

import argparse
import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

import cmcrameri.cm as cmc
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

from trustk.plotting.style import export_figure, journal_width, set_trustk_style


FT_TO_M = 0.3048

MWP1_WORKBOOK = "OFR2019_1133_DataRelease/Multi-WellPumpingTests/data/Multi-WellPumpingTest1_Data.xlsx"
MWP2_WORKBOOK = "OFR2019_1133_DataRelease/Multi-WellPumpingTests/data/Multi-WellPumpingTest2_Data.xlsx"
SINGLE_WORKBOOK = "OFR2019_1133_DataRelease/Single-WellPumpingTest/data/Single-WellPumpingTest_Data.xlsx"

MWP1_BASE = "OFR2019_1133_DataRelease/Multi-WellPumpingTests/analyses/MF_LH-MooRAH/"
MWP2_BASE = "OFR2019_1133_DataRelease/Multi-WellPumpingTests/analyses/MF_LH-Bongo/"


@dataclass(frozen=True)
class DrawdownProduct:
    test: str
    base: str
    stem: str
    pumping_well: str
    pumping_map_number: int


PRODUCTS = [
    DrawdownProduct("MWP1_MooRAH", MWP1_BASE, "LH-MooRAH", "MOIR", 4),
    DrawdownProduct("MWP2_Bongo", MWP2_BASE, "LH-Bongo", "BING", 10),
]

WELL_META = {
    "FETH": (1, "coarser alluvium"),
    "WAGT": (2, "coarser alluvium"),
    "AIRP": (3, "coarser alluvium"),
    "MOIR": (4, "coarser alluvium"),
    "MSTOK": (5, "Lahontan clays and silts"),
    "MS": (6, "Lahontan clays and silts"),
    "MD": (7, "Lahontan clays and silts"),
    "NDOA": (8, "Lahontan clays and silts"),
    "LEID": (9, "coarser alluvium"),
    "BING": (10, "coarser alluvium"),
    "WWBN": (11, "coarser alluvium"),
    "BS": (12, "Lahontan clays and silts"),
    "BD": (13, "Lahontan clays and silts"),
    "PINN": (14, "coarser alluvium"),
    "WS": (15, "Lahontan clays and silts"),
    "WD": (16, "Lahontan clays and silts"),
    "TWEL": (17, "Lahontan clays and silts"),
}

TEST_PERIODS = {
    "MWP1_MooRAH": {
        "pumping_well": "MOIR",
        "pumping_map_number": 4,
        "pump_start": pd.Timestamp("2017-03-08 13:40:00"),
        "pump_stop": pd.Timestamp("2017-03-12 13:00:00"),
        "recovery_end": pd.Timestamp("2017-04-05 00:00:00"),
        "pumping_rate_gpm": 4400.0,
        "source": "USGS OFR 2019-1133 field procedures",
    },
    "MWP2_Bongo": {
        "pumping_well": "BING",
        "pumping_map_number": 10,
        "pump_start": pd.Timestamp("2017-11-28 10:53:00"),
        "pump_stop": pd.Timestamp("2017-12-04 11:28:00"),
        "recovery_end": pd.Timestamp("2018-01-04 00:00:00"),
        "pumping_rate_gpm": 6040.0,
        "source": "USGS OFR 2019-1133 field procedures",
    },
    "Single_Well_AIRP": {
        "pumping_well": "AIRP",
        "pumping_map_number": 3,
        "pump_start": pd.Timestamp("2017-08-09 07:18:00"),
        "pump_stop": pd.Timestamp("2017-08-09 15:36:00"),
        "recovery_end": pd.Timestamp("2017-08-10 12:57:00"),
        "pumping_rate_gpm": 16.2,
        "source": "USGS OFR 2019-1133 single-well test section",
    },
}

REPORT_DISTANCE_MI = {
    "FETH": {"MWP1_MooRAH": 1.2, "MWP2_Bongo": 5.7},
    "WAGT": {"MWP1_MooRAH": 1.5, "MWP2_Bongo": 4.6},
    "AIRP": {"MWP1_MooRAH": 1.2, "MWP2_Bongo": 6.0, "Single_Well_AIRP": 0.0},
    "MOIR": {"MWP1_MooRAH": 0.0, "MWP2_Bongo": 5.0},
    "MSTOK": {"MWP1_MooRAH": 0.3},
    "MS": {"MWP2_Bongo": 4.8},
    "MD": {"MWP2_Bongo": 4.8},
    "NDOA": {"MWP1_MooRAH": 2.4, "MWP2_Bongo": 2.6},
    "LEID": {"MWP2_Bongo": 1.0},
    "BING": {"MWP1_MooRAH": 5.0, "MWP2_Bongo": 0.0},
    "WWBN": {"MWP2_Bongo": 0.6},
    "BS": {"MWP2_Bongo": 0.6},
    "BD": {"MWP2_Bongo": 0.6},
    "PINN": {"MWP2_Bongo": 5.1},
    "WS": {"MWP2_Bongo": 9.6},
    "WD": {"MWP2_Bongo": 9.6},
    "TWEL": {"MWP1_MooRAH": 16.4, "MWP2_Bongo": 12.3},
}

# Values transcribed from USGS OFR 2019-1133 tables 5 and 6.
REPORT_DRAWDOWN = {
    ("MWP1_MooRAH", "FETH"): (11.578, "High", 0.2297),
    ("MWP1_MooRAH", "WAGT"): (8.184, "High", 0.2567),
    ("MWP1_MooRAH", "AIRP"): (0.267, "High", 0.0312),
    ("MWP1_MooRAH", "NDOA"): (0.0838, "Low", 0.0118),
    ("MWP1_MooRAH", "BING"): (0.115, "Moderate", 0.0115),
    ("MWP1_MooRAH", "TWEL"): (0.0120, "Low", 5.8167),
    ("MWP2_Bongo", "FETH"): (0.126, "Moderate", 0.0113),
    ("MWP2_Bongo", "AIRP"): (0.047, "Low", 0.0037),
    ("MWP2_Bongo", "MOIR"): (0.105, "Moderate", 0.0106),
    ("MWP2_Bongo", "MS"): (0.1217, "Low", 0.0480),
    ("MWP2_Bongo", "MD"): (0.0968, "Low", 0.0396),
    ("MWP2_Bongo", "NDOA"): (0.0428, "Low", 0.0136),
    ("MWP2_Bongo", "LEID"): (2.629, "High", 0.0148),
    ("MWP2_Bongo", "WWBN"): (9.030, "High", 0.0336),
    ("MWP2_Bongo", "BS"): (0.0604, "Moderate", 0.0068),
    ("MWP2_Bongo", "BD"): (0.0615, "Moderate", 0.0063),
    ("MWP2_Bongo", "PINN"): (0.0386, "Low", 0.0041),
    ("MWP2_Bongo", "WS"): (0.0629, "Low", 0.0147),
    ("MWP2_Bongo", "WD"): (0.0840, "Low", 0.0176),
    ("MWP2_Bongo", "TWEL"): (0.0462, "Low", 0.0152),
}

USGS_FIGURE_WELLS = {
    ("MWP1_MooRAH", "FETH"),
    ("MWP1_MooRAH", "WAGT"),
    ("MWP1_MooRAH", "AIRP"),
    ("MWP1_MooRAH", "NDOA"),
    ("MWP1_MooRAH", "BING"),
    ("MWP2_Bongo", "NDOA"),
    ("MWP2_Bongo", "LEID"),
    ("MWP2_Bongo", "WWBN"),
    ("MWP2_Bongo", "BS"),
    ("MWP2_Bongo", "BD"),
}


def normalize_well_name(raw_name: str) -> str:
    name = raw_name.strip()
    name = re.sub(r"^W_", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\.ft$", "", name, flags=re.IGNORECASE)
    name = name.replace("_", "")
    if name.upper() not in WELL_META and name and name[0].lower() in {"b", "m"} and len(name) > 2:
        stripped = name[1:]
        if stripped.upper() in WELL_META:
            name = stripped
    aliases = {"AIRPORT": "AIRP", "WAGONT": "WAGT"}
    return aliases.get(name.upper(), name.upper())


def _load_workbook(archive: zipfile.ZipFile, member: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(archive.read(member)), read_only=True, data_only=True)


def _sheet_summary(workbook: openpyxl.Workbook, test: str, member: str) -> pd.DataFrame:
    rows = []
    skip = {"readme", "Barometer", "Flow_RyePatch", "Pumping", "Pumping_Data"}
    for sheet in workbook.sheetnames:
        if sheet in skip:
            continue
        ws = workbook[sheet]
        header = [str(value).strip() if value is not None else "" for value in next(ws.iter_rows(values_only=True))]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            timestamp = pd.to_datetime(row[0], errors="coerce")
            if pd.isna(timestamp):
                continue
            try:
                value = float(row[1])
            except (TypeError, ValueError):
                continue
            records.append((timestamp, value))
        if not records:
            continue
        values = np.array([value for _, value in records], dtype=float)
        well = normalize_well_name(sheet)
        map_number, aquifer = WELL_META.get(well, (np.nan, "unknown"))
        rows.append(
            {
                "test": test,
                "source_type": "raw water-level workbook",
                "source_file": member,
                "sheet": sheet,
                "well": well,
                "map_well_number": map_number,
                "primary_aquifer": aquifer,
                "n_records": len(records),
                "start": min(t for t, _ in records),
                "end": max(t for t, _ in records),
                "water_level_column": header[1] if len(header) > 1 else "",
                "water_level_min_ft": float(np.nanmin(values)),
                "water_level_max_ft": float(np.nanmax(values)),
                "water_level_range_ft": float(np.nanmax(values) - np.nanmin(values)),
            }
        )
    return pd.DataFrame(rows)


def _schedule_summary(workbook: openpyxl.Workbook, test: str, member: str) -> pd.DataFrame:
    sheet_name = "Pumping" if "Pumping" in workbook.sheetnames else "Pumping_Data"
    if sheet_name not in workbook.sheetnames:
        return pd.DataFrame()
    ws = workbook[sheet_name]
    header = [str(value).strip() if value is not None else "" for value in next(ws.iter_rows(values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        timestamp = pd.to_datetime(row[0], errors="coerce")
        if pd.isna(timestamp):
            continue
        values = {"datetime": timestamp}
        for idx, col in enumerate(header[1:], start=1):
            if row[idx] is not None:
                values[col] = row[idx]
        rows.append(values)
    if not rows:
        return pd.DataFrame()
    data = pd.DataFrame(rows)
    numeric_cols = [col for col in data.columns if col != "datetime" and pd.api.types.is_numeric_dtype(data[col])]
    flow_col = next((col for col in numeric_cols if "q_" in col.lower() or "rate" in col.lower()), numeric_cols[-1])
    flow = pd.to_numeric(data[flow_col], errors="coerce")
    return pd.DataFrame(
        [
            {
                "test": test,
                "source_type": "pumping schedule",
                "source_file": member,
                "sheet": sheet_name,
                "n_records": len(data),
                "start": data["datetime"].min(),
                "end": data["datetime"].max(),
                "flow_column": flow_col,
                "flow_min": float(flow.min()),
                "flow_max": float(flow.max()),
                "flow_median_positive": float(flow[flow > 0].median()) if (flow > 0).any() else np.nan,
            }
        ]
    )


def _parse_drawdown_text(text: str, value_name: str) -> pd.DataFrame:
    rows = []
    for line in text.splitlines():
        parts = line.split()
        if len(parts) < 4:
            continue
        timestamp = pd.to_datetime(f"{parts[1]} {parts[2]}", errors="coerce")
        if pd.isna(timestamp):
            continue
        rows.append(
            {
                "well": normalize_well_name(parts[0]),
                "datetime": timestamp,
                f"{value_name}_drawdown_ft": float(parts[3]),
                f"{value_name}_drawdown_m": float(parts[3]) * FT_TO_M,
            }
        )
    return pd.DataFrame(rows)


def _drawdown_product_summary(archive: zipfile.ZipFile, product: DrawdownProduct) -> tuple[pd.DataFrame, pd.DataFrame]:
    measured = _parse_drawdown_text(
        archive.read(f"{product.base}Pcomp_{product.stem}.WLmeasured.txt").decode("latin-1", errors="replace"),
        "measured",
    )
    simulated = _parse_drawdown_text(
        archive.read(f"{product.base}Pcomp_{product.stem}.WLsimulated.txt").decode("latin-1", errors="replace"),
        "simulated",
    )
    merged = measured.merge(simulated, on=["well", "datetime"], how="outer").sort_values(["well", "datetime"])
    rows = []
    for well, group in merged.groupby("well"):
        report = REPORT_DRAWDOWN.get((product.test, well), (np.nan, "not listed", np.nan))
        map_number, aquifer = WELL_META.get(well, (np.nan, "unknown"))
        role = "pumping well" if well == product.pumping_well else "observation well"
        listed = bool(np.isfinite(report[0]))
        measured_values = group["measured_drawdown_ft"].to_numpy(float)
        simulated_values = group["simulated_drawdown_ft"].to_numpy(float)
        residual = measured_values - simulated_values
        rows.append(
            {
                "test": product.test,
                "source_type": "USGS water-level-model drawdown product",
                "well": well,
                "map_well_number": map_number,
                "role": role,
                "primary_aquifer": aquifer,
                "n_drawdown_points": len(group),
                "start": group["datetime"].min(),
                "end": group["datetime"].max(),
                "max_measured_drawdown_ft": float(np.nanmax(measured_values)),
                "min_measured_drawdown_ft": float(np.nanmin(measured_values)),
                "max_abs_measured_drawdown_ft": float(np.nanmax(np.abs(measured_values))),
                "max_simulated_drawdown_ft": float(np.nanmax(simulated_values)),
                "rms_residual_ft": float(np.sqrt(np.nanmean(residual**2))),
                "report_estimated_drawdown_ft": report[0],
                "report_relative_certainty": report[1],
                "report_rms_error_ft": report[2],
                "above_detection_threshold_0p05ft": bool((report[0] if listed else np.nanmax(measured_values)) >= 0.05),
                "listed_in_usgs_table": listed,
                "shown_in_usgs_overview_figure": (product.test, well) in USGS_FIGURE_WELLS,
                "recommended_for_clear_pumping_curve": bool(listed and report[1] in {"High", "Moderate"} and report[0] >= 0.05),
                "notes": _drawdown_notes(product, well, report[0], report[1]),
            }
        )
    timeseries = merged.copy()
    timeseries["test"] = product.test
    period = TEST_PERIODS[product.test]
    timeseries["phase"] = timeseries["datetime"].map(lambda dt: _phase_for_datetime(dt, period))
    timeseries["elapsed_since_pump_start_days"] = (
        timeseries["datetime"] - period["pump_start"]
    ).dt.total_seconds() / 86400.0
    timeseries["elapsed_since_pump_stop_days"] = (
        timeseries["datetime"] - period["pump_stop"]
    ).dt.total_seconds() / 86400.0
    return pd.DataFrame(rows), timeseries


def _drawdown_notes(product: DrawdownProduct, well: str, estimated_drawdown: float, certainty: str) -> str:
    if well == product.pumping_well:
        return "pumping well; report says pumping well drawdown was not evaluated as an observation well"
    if not np.isfinite(estimated_drawdown):
        return "not listed in USGS report drawdown table"
    if estimated_drawdown < 0.05:
        return "below the 0.05 ft drawdown-detection threshold used by USGS"
    if certainty == "Low":
        return "above 0.05 ft but classified by USGS as low-certainty"
    return "usable report drawdown response"


def _phase_for_datetime(timestamp: pd.Timestamp, period: dict) -> str:
    if timestamp < period["pump_start"]:
        return "pre-pumping"
    if timestamp <= period["pump_stop"]:
        return "pumping"
    if timestamp <= period["recovery_end"]:
        return "recovery"
    return "post-recovery"


def _build_period_table() -> pd.DataFrame:
    rows = []
    for test, period in TEST_PERIODS.items():
        duration_h = (period["pump_stop"] - period["pump_start"]).total_seconds() / 3600.0
        recovery_d = (period["recovery_end"] - period["pump_stop"]).total_seconds() / 86400.0
        rows.append(
            {
                "test": test,
                "pumping_well": period["pumping_well"],
                "pumping_map_number": period["pumping_map_number"],
                "pump_start": period["pump_start"],
                "pump_stop": period["pump_stop"],
                "pumping_duration_hours": duration_h,
                "recovery_end": period["recovery_end"],
                "recovery_monitoring_days": recovery_d,
                "pumping_rate_gpm": period["pumping_rate_gpm"],
                "source": period["source"],
            }
        )
    return pd.DataFrame(rows)


def _build_distance_table(drawdown: pd.DataFrame, raw: pd.DataFrame) -> pd.DataFrame:
    raw_keys = {(row.test, row.well) for row in raw.itertuples(index=False)}
    drawdown_lookup = {
        (row.test, row.well): row
        for row in drawdown.itertuples(index=False)
    }
    rows = []
    for well, distances in REPORT_DISTANCE_MI.items():
        map_number, aquifer = WELL_META.get(well, (np.nan, "unknown"))
        for test, distance_mi in distances.items():
            period = TEST_PERIODS[test]
            key = (test, well)
            report = REPORT_DRAWDOWN.get(key, (np.nan, "", np.nan))
            role = "pumping well" if well == period["pumping_well"] else "observation well"
            if key in drawdown_lookup:
                data_status = "USGS drawdown product"
            elif key in raw_keys:
                data_status = "raw water level only"
            else:
                data_status = "not evaluated or not in extracted workbook"
            rows.append(
                {
                    "test": test,
                    "pumping_well": period["pumping_well"],
                    "pumping_map_number": period["pumping_map_number"],
                    "well": well,
                    "map_well_number": map_number,
                    "role": role,
                    "primary_aquifer": aquifer,
                    "distance_mi_report": distance_mi,
                    "distance_km_report": distance_mi * 1.609344,
                    "report_estimated_drawdown_ft": report[0],
                    "report_relative_certainty": report[1],
                    "report_rms_error_ft": report[2],
                    "data_status": data_status,
                }
            )
    return pd.DataFrame(rows).sort_values(["test", "distance_mi_report", "map_well_number"]).reset_index(drop=True)


def _build_phase_summary(timeseries: pd.DataFrame, distance: pd.DataFrame) -> pd.DataFrame:
    summary = (
        timeseries.groupby(["test", "well", "phase"], as_index=False)
        .agg(
            n_points=("measured_drawdown_ft", "size"),
            start=("datetime", "min"),
            end=("datetime", "max"),
            max_measured_drawdown_ft=("measured_drawdown_ft", "max"),
            min_measured_drawdown_ft=("measured_drawdown_ft", "min"),
        )
    )
    keep = distance[["test", "well", "map_well_number", "distance_mi_report", "distance_km_report"]]
    return summary.merge(keep, on=["test", "well"], how="left").sort_values(
        ["test", "map_well_number", "phase"]
    )


def build_inventory(zip_path: str | Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    zip_file = Path(zip_path)
    raw_tables = []
    schedules = []
    drawdown_tables = []
    drawdown_timeseries = []
    with zipfile.ZipFile(zip_file) as archive:
        for test, member in [
            ("MWP1_MooRAH", MWP1_WORKBOOK),
            ("MWP2_Bongo", MWP2_WORKBOOK),
            ("Single_Well_AIRP", SINGLE_WORKBOOK),
        ]:
            workbook = _load_workbook(archive, member)
            raw_tables.append(_sheet_summary(workbook, test, member))
            schedule = _schedule_summary(workbook, test, member)
            if not schedule.empty:
                schedules.append(schedule)
        for product in PRODUCTS:
            summary, timeseries = _drawdown_product_summary(archive, product)
            drawdown_tables.append(summary)
            drawdown_timeseries.append(timeseries)
    return (
        pd.concat(raw_tables, ignore_index=True),
        pd.concat(schedules, ignore_index=True),
        pd.concat(drawdown_tables, ignore_index=True),
        pd.concat(drawdown_timeseries, ignore_index=True),
    )


def write_inventory_outputs(
    raw: pd.DataFrame,
    schedules: pd.DataFrame,
    drawdown: pd.DataFrame,
    timeseries: pd.DataFrame,
    out_dir: str | Path,
) -> None:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    periods = _build_period_table()
    distances = _build_distance_table(drawdown, raw)
    phase_summary = _build_phase_summary(timeseries, distances)
    raw.to_csv(out / "usgs_pumping_raw_water_level_inventory.csv", index=False)
    schedules.to_csv(out / "usgs_pumping_schedule_inventory.csv", index=False)
    drawdown.to_csv(out / "usgs_pumping_drawdown_inventory.csv", index=False)
    timeseries.to_csv(out / "usgs_pumping_drawdown_timeseries.csv", index=False)
    periods.to_csv(out / "usgs_pumping_test_periods.csv", index=False)
    distances.to_csv(out / "usgs_pumping_well_distances.csv", index=False)
    phase_summary.to_csv(out / "usgs_pumping_phase_summary_by_well.csv", index=False)
    _write_markdown_summary(raw, schedules, drawdown, periods, distances, out / "usgs_pumping_inventory_summary.md")


def _write_markdown_summary(
    raw: pd.DataFrame,
    schedules: pd.DataFrame,
    drawdown: pd.DataFrame,
    periods: pd.DataFrame,
    distances: pd.DataFrame,
    out_file: Path,
) -> None:
    lines = [
        "# USGS Lovelock Pumping-Test Inventory",
        "",
        "This file separates raw transducer water levels from USGS water-level-model drawdown products.",
        "",
        "## Pumping and recovery periods",
        "",
        periods.to_markdown(index=False),
        "",
        "## Report distances from pumping wells",
        "",
        distances[
            [
                "test",
                "pumping_well",
                "well",
                "map_well_number",
                "role",
                "distance_mi_report",
                "distance_km_report",
                "report_estimated_drawdown_ft",
                "report_relative_certainty",
                "data_status",
            ]
        ].to_markdown(index=False),
        "",
        "## Pumping schedules",
        "",
        schedules.to_markdown(index=False),
        "",
        "## Raw water-level sheets",
        "",
        raw[["test", "well", "map_well_number", "n_records", "start", "end", "water_level_range_ft"]]
        .sort_values(["test", "map_well_number"])
        .to_markdown(index=False),
        "",
        "## USGS drawdown products",
        "",
        drawdown[
            [
                "test",
                "well",
                "map_well_number",
                "role",
                "n_drawdown_points",
                "report_estimated_drawdown_ft",
                "report_relative_certainty",
                "above_detection_threshold_0p05ft",
                "shown_in_usgs_overview_figure",
                "recommended_for_clear_pumping_curve",
                "notes",
            ]
        ]
        .sort_values(["test", "map_well_number"])
        .to_markdown(index=False),
        "",
    ]
    out_file.write_text("\n".join(lines), encoding="utf-8")


def plot_inventory(drawdown: pd.DataFrame, timeseries: pd.DataFrame, out_dir: str | Path) -> None:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    set_trustk_style()
    fig, axes = plt.subplots(2, 2, figsize=(journal_width(190), 6.9), gridspec_kw={"height_ratios": [1.05, 1.0]})

    _plot_report_bars(axes[0, 0], drawdown, "MWP1_MooRAH")
    _plot_report_bars(axes[0, 1], drawdown, "MWP2_Bongo")
    _plot_time_series(axes[1, 0], drawdown, timeseries, "MWP1_MooRAH")
    _plot_time_series(axes[1, 1], drawdown, timeseries, "MWP2_Bongo")

    for label, ax in zip(["(a)", "(b)", "(c)", "(d)"], axes.ravel()):
        ax.text(0.02, 0.96, label, transform=ax.transAxes, ha="left", va="top", fontsize=9, fontweight="bold")
    fig.subplots_adjust(left=0.070, right=0.985, bottom=0.155, top=0.95, wspace=0.30, hspace=0.58)
    export_figure(fig, out / "diagnostic_usgs_pumping_inventory")
    plt.close(fig)
    _plot_phase_distance(drawdown, timeseries, _build_distance_table(drawdown, pd.DataFrame()), out)


def _certainty_color(certainty: str) -> tuple[float, float, float, float]:
    return {
        "High": cmc.batlow(0.82),
        "Moderate": cmc.batlow(0.55),
        "Low": cmc.batlow(0.25),
        "not listed": (0.76, 0.76, 0.76, 1.0),
    }.get(str(certainty), (0.76, 0.76, 0.76, 1.0))


def _plot_report_bars(ax, drawdown: pd.DataFrame, test: str) -> None:
    data = drawdown[(drawdown["test"].eq(test)) & drawdown["listed_in_usgs_table"]].copy()
    data = data.sort_values("map_well_number")
    x = np.arange(len(data))
    y = data["report_estimated_drawdown_ft"].to_numpy(float)
    colors = [_certainty_color(cert) for cert in data["report_relative_certainty"]]
    ax.bar(x, y, color=colors, edgecolor="0.2", linewidth=0.35)
    ax.axhline(0.05, color="0.2", lw=0.8, ls=":", label="0.05 ft threshold")
    ax.set_yscale("log")
    ax.set_xticks(x)
    labels = [f"{int(row.map_well_number)} {row.well}" for row in data.itertuples(index=False)]
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=6.0)
    ax.set_ylabel("USGS estimated drawdown (ft)")
    ax.set_title(test.replace("_", " "), fontsize=8, pad=4)
    ax.spines[["top", "right"]].set_visible(False)
    handles = [
        plt.Line2D([0], [0], marker="s", linestyle="", color=_certainty_color("High"), label="High"),
        plt.Line2D([0], [0], marker="s", linestyle="", color=_certainty_color("Moderate"), label="Moderate"),
        plt.Line2D([0], [0], marker="s", linestyle="", color=_certainty_color("Low"), label="Low"),
        plt.Line2D([0], [0], color="0.2", ls=":", label="0.05 ft"),
    ]
    ax.legend(handles=handles, frameon=False, loc="upper right", ncol=2, columnspacing=0.6, handlelength=1.2)


def _plot_time_series(ax, drawdown: pd.DataFrame, timeseries: pd.DataFrame, test: str) -> None:
    candidates = drawdown[
        drawdown["test"].eq(test)
        & drawdown["shown_in_usgs_overview_figure"]
        & drawdown["recommended_for_clear_pumping_curve"]
    ].sort_values("map_well_number")
    if candidates.empty:
        candidates = drawdown[drawdown["test"].eq(test) & drawdown["shown_in_usgs_overview_figure"]].sort_values(
            "map_well_number"
        )
    selected = candidates["well"].tolist()
    colors = {well: cmc.vik(value) for well, value in zip(selected, np.linspace(0.15, 0.88, max(len(selected), 2)))}
    for well in selected:
        group = timeseries[(timeseries["test"].eq(test)) & (timeseries["well"].eq(well))].sort_values("datetime")
        if group.empty:
            continue
        elapsed_days = (group["datetime"] - group["datetime"].min()).dt.total_seconds() / 86400.0
        y = group["measured_drawdown_ft"].to_numpy(float)
        ax.plot(elapsed_days, y, marker="o", ms=2.0, lw=0.8, color=colors[well], label=well)
    ax.axhline(0.05, color="0.2", lw=0.75, ls=":")
    ax.set_xlabel("Elapsed time from first drawdown estimate (days)")
    ax.set_ylabel("USGS estimated drawdown (ft)")
    ax.spines[["top", "right"]].set_visible(False)
    ax.legend(frameon=False, loc="best", ncol=2, handlelength=1.2)
    ax.set_title(test.replace("_", " ") + " selected report curves", fontsize=8, pad=4)


def _plot_phase_distance(drawdown: pd.DataFrame, timeseries: pd.DataFrame, distance: pd.DataFrame, out: Path) -> None:
    set_trustk_style()
    fig, axes = plt.subplots(1, 2, figsize=(journal_width(190), 3.45))
    for ax, test in zip(axes, ["MWP1_MooRAH", "MWP2_Bongo"]):
        period = TEST_PERIODS[test]
        candidates = drawdown[
            drawdown["test"].eq(test)
            & drawdown["listed_in_usgs_table"]
            & drawdown["recommended_for_clear_pumping_curve"]
        ].copy()
        candidates = candidates.sort_values(["report_estimated_drawdown_ft"], ascending=False)
        selected = candidates["well"].tolist()[:5]
        colors = {well: cmc.batlow(value) for well, value in zip(selected, np.linspace(0.15, 0.85, max(len(selected), 2)))}
        pump_stop_day = (period["pump_stop"] - period["pump_start"]).total_seconds() / 86400.0
        ax.axvspan(0, pump_stop_day, color="0.92", zorder=0, label="pumping")
        ax.axvline(0, color="0.25", lw=0.75, ls=":")
        ax.axvline(pump_stop_day, color="0.25", lw=0.85, ls="--", label="pump off")
        for well in selected:
            group = timeseries[(timeseries["test"].eq(test)) & (timeseries["well"].eq(well))].sort_values("datetime")
            dist = distance[(distance["test"].eq(test)) & (distance["well"].eq(well))]
            dist_label = ""
            if not dist.empty:
                dist_label = f", {float(dist['distance_mi_report'].iloc[0]):.1f} mi"
            ax.plot(
                group["elapsed_since_pump_start_days"],
                group["measured_drawdown_ft"],
                marker="o",
                ms=2.0,
                lw=0.9,
                color=colors[well],
                label=f"{well}{dist_label}",
            )
        ax.set_xlabel("Elapsed time since pump start (days)")
        ax.set_ylabel("USGS estimated drawdown (ft)")
        ax.set_title(f"{test.replace('_', ' ')}: pumping then recovery", fontsize=8, pad=4)
        ax.spines[["top", "right"]].set_visible(False)
        ax.legend(frameon=False, loc="best", fontsize=6.6, handlelength=1.4)
    fig.subplots_adjust(left=0.065, right=0.985, bottom=0.17, top=0.88, wspace=0.28)
    export_figure(fig, out / "diagnostic_usgs_pumping_phases_distances")
    plt.close(fig)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--zip", default="field data/OFR2019_1133_DataRelease.zip")
    parser.add_argument("--table-dir", default="outputs/tables")
    parser.add_argument("--figure-dir", default="outputs/figures")
    args = parser.parse_args(argv)
    raw, schedules, drawdown, timeseries = build_inventory(args.zip)
    write_inventory_outputs(raw, schedules, drawdown, timeseries, args.table_dir)
    plot_inventory(drawdown, timeseries, args.figure_dir)
    print(f"raw_water_level_rows={len(raw)}")
    print(f"schedule_rows={len(schedules)}")
    print(f"drawdown_summary_rows={len(drawdown)}")
    print(f"drawdown_timeseries_rows={len(timeseries)}")
    print(f"wrote_tables={args.table_dir}")
    print(f"wrote_figure={Path(args.figure_dir, 'diagnostic_usgs_pumping_inventory.pdf')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
