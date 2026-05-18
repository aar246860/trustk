"""Extract pumping and recovery records from USGS Lovelock multi-well test #2."""

from __future__ import annotations

import argparse
import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

PUMPING_TEST2_WORKBOOK = (
    "OFR2019_1133_DataRelease/Multi-WellPumpingTests/data/Multi-WellPumpingTest2_Data.xlsx"
)

BONGO_COORDS = (
    "OFR2019_1133_DataRelease/Multi-WellPumpingTests/analyses/MF_LH-Bongo/"
    "Pcomp_LH-Bongo.ObsWellCoord.txt"
)

OVERLAP_WELLS = {"BD", "BS", "MD", "MS", "WD", "WS", "PINN"}
KNOWN_WELLS = {
    "AIRP",
    "BD",
    "BING",
    "BS",
    "FETH",
    "LEID",
    "MD",
    "MOIR",
    "MS",
    "NDOA",
    "PINN",
    "TWEL",
    "WAGT",
    "WD",
    "WS",
    "WWBN",
}
FT_TO_M = 0.3048
GPM_TO_M3_S = 0.0000630901964


@dataclass(frozen=True)
class PumpingExtraction:
    schedule: pd.DataFrame
    timeseries: pd.DataFrame
    coordinates: pd.DataFrame


def _load_workbook_from_zip(archive: zipfile.ZipFile, member: str) -> openpyxl.Workbook:
    data = archive.read(member)
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def normalize_well_name(raw_name: str) -> str:
    """Normalize USGS/PEST well names such as bBD__ or W_BD.ft to BD."""

    name = raw_name.strip()
    name = re.sub(r"^W_", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\.ft$", "", name, flags=re.IGNORECASE)
    name = name.replace("_", "")
    if name.upper() not in KNOWN_WELLS and name and name[0].lower() in {"b", "m"} and len(name) > 2:
        stripped = name[1:]
        if stripped.upper() in KNOWN_WELLS:
            name = stripped
    aliases = {"AIRPORT": "AIRP", "WAGONT": "WAGT"}
    return aliases.get(name.upper(), name.upper())


def _extract_schedule(wb: openpyxl.Workbook) -> pd.DataFrame:
    ws = wb["Pumping"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        q_gpm = float(row[1])
        rows.append(
            {
                "datetime": row[0],
                "q_gpm": q_gpm,
                "q_m3_s": q_gpm * GPM_TO_M3_S,
            }
        )
    schedule = pd.DataFrame(rows)
    schedule["datetime"] = pd.to_datetime(schedule["datetime"])
    return schedule.sort_values("datetime").reset_index(drop=True)


def _phase_for_time(t: pd.Timestamp, pump_start: pd.Timestamp, pump_stop: pd.Timestamp) -> str:
    if t < pump_start:
        return "pretest"
    if t <= pump_stop:
        return "pumping"
    return "recovery"


def _extract_timeseries(wb: openpyxl.Workbook, schedule: pd.DataFrame) -> pd.DataFrame:
    pump_start = schedule.loc[schedule["q_gpm"] > 0, "datetime"].min()
    pump_stop = schedule.loc[schedule["q_gpm"] > 0, "datetime"].max()
    rows = []
    for sheet in wb.sheetnames:
        if sheet in {"readme", "Pumping", "Barometer"}:
            continue
        ws = wb[sheet]
        header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(values_only=True))]
        if len(header) < 2:
            continue
        value_col = header[1]
        well = normalize_well_name(sheet)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            dt = pd.Timestamp(row[0])
            water_level_ft = float(row[1])
            rows.append(
                {
                    "well": well,
                    "datetime": dt,
                    "water_level_ft": water_level_ft,
                    "water_level_m": water_level_ft * FT_TO_M,
                    "phase": _phase_for_time(dt, pump_start, pump_stop),
                    "is_slug_overlap_well": well in OVERLAP_WELLS,
                    "source_column": value_col,
                }
            )
    data = pd.DataFrame(rows).sort_values(["well", "datetime"]).reset_index(drop=True)
    data["elapsed_since_pump_start_s"] = (
        data["datetime"] - pump_start
    ).dt.total_seconds()
    data["elapsed_since_recovery_start_s"] = (
        data["datetime"] - pump_stop
    ).dt.total_seconds()
    return data


def _extract_coordinates(archive: zipfile.ZipFile) -> pd.DataFrame:
    text = archive.read(BONGO_COORDS).decode("latin-1", errors="replace")
    rows = []
    for line in text.splitlines():
        parts = line.split()
        if len(parts) < 4:
            continue
        rows.append(
            {
                "well": normalize_well_name(parts[0]),
                "x_m": float(parts[1]),
                "y_m": float(parts[2]),
                "model_layer": int(float(parts[3])),
                "source_file": BONGO_COORDS,
            }
        )
    return pd.DataFrame(rows).sort_values("well").reset_index(drop=True)


def extract_pumping_test2(zip_path: str | Path) -> PumpingExtraction:
    """Extract multi-well pumping test #2 schedule, observations, and coordinates."""

    with zipfile.ZipFile(zip_path) as archive:
        wb = _load_workbook_from_zip(archive, PUMPING_TEST2_WORKBOOK)
        schedule = _extract_schedule(wb)
        timeseries = _extract_timeseries(wb, schedule)
        coordinates = _extract_coordinates(archive)
    return PumpingExtraction(schedule=schedule, timeseries=timeseries, coordinates=coordinates)


def write_pumping_outputs(extraction: PumpingExtraction, out_dir: str | Path) -> None:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    extraction.schedule.to_csv(out_dir / "pumping_test2_schedule.csv", index=False)
    extraction.timeseries.to_csv(out_dir / "pumping_test2_timeseries.csv", index=False)
    extraction.coordinates.to_csv(out_dir / "pumping_test2_well_coordinates.csv", index=False)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--zip", required=True, help="Path to OFR2019_1133_DataRelease.zip")
    parser.add_argument("--out-dir", default="data/processed", help="Output directory")
    args = parser.parse_args(argv)
    extraction = extract_pumping_test2(args.zip)
    write_pumping_outputs(extraction, args.out_dir)
    print(f"pumping_schedule_rows={len(extraction.schedule)}")
    print(f"pumping_timeseries_rows={len(extraction.timeseries)}")
    print(f"well_coordinates={len(extraction.coordinates)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
