"""Extract slug-test metadata and curves from the USGS Lovelock data release."""

from __future__ import annotations

import argparse
import io
import zipfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

SLUG_ANALYSIS_WORKBOOKS = {
    "PINN": "OFR2019_1133_DataRelease/InjectionSlugTest_YoungerAlluviumCoarserSediments/analyses/400719118305301_BouwerRice_SlugTest.xlsm",
    "BD": "OFR2019_1133_DataRelease/InjectionSlugTests_YoungerAlluviumFinerSediments/analyses/BD_BouwerRice_SlugTest_DTW.xlsm",
    "BS": "OFR2019_1133_DataRelease/InjectionSlugTests_YoungerAlluviumFinerSediments/analyses/BS_BouwerRice_SlugTest_DTW.xlsm",
    "MD": "OFR2019_1133_DataRelease/InjectionSlugTests_YoungerAlluviumFinerSediments/analyses/MD_BouwerRice_SlugTest_DTW.xlsm",
    "MS": "OFR2019_1133_DataRelease/InjectionSlugTests_YoungerAlluviumFinerSediments/analyses/MS_BouwerRice_SlugTest_DTW.xlsm",
    "WD": "OFR2019_1133_DataRelease/InjectionSlugTests_YoungerAlluviumFinerSediments/analyses/WD_BouwerRice_SlugTest_DTW.xlsm",
    "WS": "OFR2019_1133_DataRelease/InjectionSlugTests_YoungerAlluviumFinerSediments/analyses/WS_BouwerRice_SlugTest_DTW.xlsm",
}

FT_TO_M = 0.3048


@dataclass(frozen=True)
class SlugExtraction:
    tests: pd.DataFrame
    curves: pd.DataFrame


def _load_workbook_from_zip(archive: zipfile.ZipFile, member: str) -> openpyxl.Workbook:
    data = archive.read(member)
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def _cell(ws: openpyxl.worksheet.worksheet.Worksheet, address: str):
    return ws[address].value


def _extract_curve(wb: openpyxl.Workbook, well: str, source_file: str) -> pd.DataFrame:
    ws = wb["Drawdown"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rows.append(
            {
                "well": well,
                "datetime": row[0],
                "depth_to_water_ft": row[1],
                "drawdown_ft": row[2],
                "source_file": source_file,
            }
        )
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df["datetime"] = pd.to_datetime(df["datetime"])
    start = df["datetime"].min()
    df["elapsed_s"] = (df["datetime"] - start).dt.total_seconds()
    df["drawdown_m"] = df["drawdown_ft"] * FT_TO_M
    df["depth_to_water_m"] = df["depth_to_water_ft"] * FT_TO_M
    return df


def extract_slug_tests(zip_path: str | Path) -> SlugExtraction:
    """Extract Bouwer-Rice slug-test summaries and curves."""

    tests: list[dict] = []
    curves: list[pd.DataFrame] = []
    with zipfile.ZipFile(zip_path) as archive:
        for well, member in SLUG_ANALYSIS_WORKBOOKS.items():
            wb = _load_workbook_from_zip(archive, member)
            comp = wb["COMPUTATION"]
            output = wb["OUTPUT"]
            k_ft_day = float(_cell(comp, "B29"))
            row = {
                "well": well,
                "site_id": str(_cell(output, "B28")),
                "well_id": str(_cell(output, "B27")),
                "k_slug_ft_day": k_ft_day,
                "k_slug_m_s": k_ft_day * FT_TO_M / 86400.0,
                "static_water_level_ft": _cell(output, "F26"),
                "casing_diameter_in": _cell(output, "B30"),
                "rw_ft": _cell(comp, "B20"),
                "rw_m": _cell(comp, "B20") * FT_TO_M,
                "wetted_length_ft": _cell(comp, "B24"),
                "wetted_length_m": _cell(comp, "B24") * FT_TO_M,
                "source_file": member,
            }
            tests.append(row)
            curves.append(_extract_curve(wb, well, member))

    tests_df = pd.DataFrame(tests).sort_values("well").reset_index(drop=True)
    curves_df = pd.concat(curves, ignore_index=True).sort_values(["well", "elapsed_s"])
    return SlugExtraction(tests=tests_df, curves=curves_df.reset_index(drop=True))


def write_slug_outputs(extraction: SlugExtraction, out_dir: str | Path) -> None:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    extraction.tests.to_csv(out_dir / "slug_tests.csv", index=False)
    extraction.curves.to_csv(out_dir / "slug_curves.csv", index=False)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--zip", required=True, help="Path to OFR2019_1133_DataRelease.zip")
    parser.add_argument("--out-dir", default="data/processed", help="Output directory")
    args = parser.parse_args(argv)
    extraction = extract_slug_tests(args.zip)
    write_slug_outputs(extraction, args.out_dir)
    print(f"slug_tests={len(extraction.tests)}")
    print(f"slug_curve_points={len(extraction.curves)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

